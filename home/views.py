from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.contrib.auth import logout
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.utils import translation
from django.shortcuts import render
from django.http import Http404
from django.db.utils import ProgrammingError, OperationalError


import logging

logger = logging.getLogger(__name__)


def custom_404(request, exception):
    return render(request, "404.html", status=404)


# def home(request):
#     return render(request, "home/home.html")


def home_view(request):
    if request.GET.get("debug_db") == "1":
        from django.db import connection
        from django.http import JsonResponse
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT table_name 
                FROM information_schema.tables 
                WHERE table_schema='public'
            """)
            tables = [row[0] for row in cursor.fetchall()]
            
            cursor.execute("SELECT app, name, applied FROM django_migrations ORDER BY applied DESC")
            migrations = [f"{row[0]}: {row[1]} ({row[2]})" for row in cursor.fetchall()]
            
            artist_cols = []
            try:
                cursor.execute("""
                    SELECT column_name, data_type 
                    FROM information_schema.columns 
                    WHERE table_name = 'home_artist'
                """)
                artist_cols = [f"{row[0]} ({row[1]})" for row in cursor.fetchall()]
            except Exception as e:
                artist_cols = [str(e)]
                
        return JsonResponse({
            "tables": tables,
            "migrations": migrations,
            "artist_columns": artist_cols
        })

    current_lang = translation.get_language()
    print(f"\n{'='*60}")
    print(f"🏠 HOME VIEW DEBUG")
    print(f"{'='*60}")
    print(f"Current Language: {current_lang}")
    print(f"Session Language: {request.session.get('django_language', 'Not Set')}")
    print(f"Request Path: {request.path}")
    print(f"Accept-Language Header: {request.META.get('HTTP_ACCEPT_LANGUAGE', 'None')}")
    logger.info(f"Home view accessed with language: {current_lang}")
    print(f"{'='*60}\n")
    return render(request, "home.html")



def about(request):
    return render(request, "about.html")


def contact_us(request):
    return render(request, "contact_us.html")


def explore(request):
    return render(request, "exlore.html")


from django.shortcuts import render


def edit_profile(request):
    return render(request, "edit_profile.html")


def book_service(request, service_name):
    # service_name will be "3d-art", "mural", or "normal-paint"
    return render(request, "book_service.html", {"service_name": service_name})


from django.shortcuts import render
from employee.models import ServiceImage  # Make sure to import your model

# ... (imports)
from django.db.models import Q


# views.py
def explore_service(request, service_type):
    service_dict = {
        "3d-wall-art": "3D Wall Art",
        "3d-floor-art": "3D Floor Art",
        "mural-art": "Mural Art",
        "mural": "Mural Art",
        # "pillere-painting": "Pillere Art",
        "metro-advertisement": "Metro Advertisement",
        "outdoor-advertisement": "Outdoor Advertisement",
        "school-painting": "School Painting",
        "selfie-painting": "Selfie Painting",
        "madhubani-painting": "Madhubani Painting",
        "texture-painting": "Texture Painting",
        "stone-murti": "Stone Murti",
        "statue": "Stone Murti",
        "Zenith-Collection": "Zenith Collection",
        "scrap-animal-art": "Scrap Animal Art",
        "nature-fountain": "Nature & Water Fountain",
        "fountain-art": "Nature & Water Fountain",
        "cartoon-painting": "Cartoon Painting",
        "home-painting": "Home Painting",
    }

    service_name = service_dict.get(service_type, "Service")
    query_term = service_name

    if service_name == "Mural Art":
        query_term = "Mural"
    elif service_name == "Nature & Water Fountain":
        query_term = "Fountain"

    # ✅ Updated logic
    if request.user.is_authenticated and request.user.is_staff:
        # Admins see everything
        db_images = ServiceImage.objects.filter(
            type_of_art__icontains=query_term, is_approved=True
        )
    elif request.user.is_authenticated:
        # Employees see their own + verified ones
        db_images = ServiceImage.objects.filter(
            type_of_art__icontains=query_term
        ).filter(
            Q(is_verified_pic=True) | Q(userupload_id=request.user.id), is_approved=True
        )
    else:
        # Guests see only verified ones
        db_images = ServiceImage.objects.filter(
            type_of_art__icontains=query_term, is_verified_pic=True, is_approved=True
        )
    return render(
        request,
        "explore_service.html",
        {
            "service_name": service_name,
            "db_images": db_images,
            "service_slug": service_type,
        },
    )


@csrf_exempt
def approve_service_image(request):
    if request.method == "POST":

        if not request.user.is_authenticated or not request.user.is_staff:
            return JsonResponse({"success": False, "message": "Permission denied."})

        try:
            data = json.loads(request.body)
            image_id = data.get("id")
            image = ServiceImage.objects.get(id=image_id)
            image.is_approved = True
            image.is_verified_pic = True
            image.save()
            return JsonResponse(
                {"success": True, "message": "Image approved successfully!"}
            )
        except ServiceImage.DoesNotExist:
            return JsonResponse({"success": False, "message": "Image not found."})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)})

    return JsonResponse({"success": False, "message": "Invalid request."})


from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json
from employee.models import ServiceImage


@csrf_exempt
def delete_service_image(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            image_id = data.get("id")
            image = ServiceImage.objects.get(id=image_id)

            user = request.user

            # Admin can delete any image
            if user.is_staff or image.userupload_id == user.id:
                if not user.is_staff and image.is_approved and image.is_verified_pic:
                    return JsonResponse(
                        {
                            "success": False,
                            "message": "Approved and verified uploads cannot be deleted.",
                        }
                    )
                image.delete()
                return JsonResponse({"success": True})
            else:
                return JsonResponse({"success": False, "message": "Permission denied."})

        except ServiceImage.DoesNotExist:
            return JsonResponse({"success": False, "message": "Image not found."})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)})

    return JsonResponse({"success": False, "message": "Invalid request."})


# ... (imports)
def book_service(request, service_type):
    """
    service_type: could be any of the defined service keys.
    """
    service_dict = {
        "3d-wall-art": "3D Wall Art",
        "3d-floor-art": "3D Floor Art",
        "mural-art": "Mural Art",
        "school-painting": "School Painting",
        "mural": "Mural Art",
        "Zenith-Collection": "Zenith Collection",
        "cartoon-painting": "Cartoon Painting",
        "metro-advertisement": "Metro Advertisement",
        "outdoor-advertisement": "Outdoor Advertisement",
        # "pillere-painting": "Pillere Art",
        "selfie-painting": "Selfie Painting",
        "madhubani-painting": "Madhubani Painting",
        "texture-painting": "Texture Painting",
        "stone-murti": "Stone Murti",
        "statue": "Stone Murti",
        "scrap-animal-art": "Scrap Animal Art",
        "nature-fountain": "Nature & Water Fountain",
        "fountain-art": "Nature & Water Fountain",
        "home-painting": "Home Painting",
        # Mapping for any previous, short slugs (optional, but good for backward compatibility)
        "3d-art": "3D Wall Art",
        "advertisement-art": "Metro Advertisement",  # Guessing based on common short form
        "aesthetic-art": "Outdoor Advertisement",  # Guessing based on common short form
        "madhubani-art": "Madhubani Painting",  # Guessing based on common short form
        "cartoon-art": "Cartoon Painting",  # Guessing based on common short form
        "nature-art": "Nature & Water Fountain",  # Guessing based on common short form
        "scrap-yard-art": "Scrap Animal Art",  # Guessing based on common short form
        "spray-art": "Statue",  # Guessing based on context
        "structure-art": "Scrap Animal Art",  # Guessing based on context
    }

    service_name = service_dict.get(service_type, "Service")

    if request.user.is_authenticated and request.user.role == "employee":
        messages.error(
            request,
            "Artists/Employees are not allowed to create bookings.",
        )
        return redirect("explore_service", service_type=service_type)

    # Use robust filtering logic similar to explore_service
    query_term = service_name

    if service_name == "Mural Art":
        query_term = "Mural"
    elif service_name == "Nature & Water Fountain":
        query_term = "Fountain"

    # Strict visibility for booking design selector:
    # show only admin-approved + verified images.
    db_images = ServiceImage.objects.filter(
        type_of_art__icontains=query_term, is_approved=True, is_verified_pic=True
    )

    if request.method == "POST":
        # Handle booking form submission if needed
        pass

    return render(
        request,
        "book_service.html",
        {"service_name": service_name, "db_images": db_images},
    )


def home(request):
    if request.GET.get("debug_db") == "1":
        from django.db import connection
        from django.http import JsonResponse
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT table_name 
                FROM information_schema.tables 
                WHERE table_schema='public'
            """)
            tables = [row[0] for row in cursor.fetchall()]
            
            cursor.execute("SELECT app, name, applied FROM django_migrations ORDER BY applied DESC")
            migrations = [f"{row[0]}: {row[1]} ({row[2]})" for row in cursor.fetchall()]
            
            artist_cols = []
            try:
                cursor.execute("""
                    SELECT column_name, data_type 
                    FROM information_schema.columns 
                    WHERE table_name = 'home_artist'
                """)
                artist_cols = [f"{row[0]} ({row[1]})" for row in cursor.fetchall()]
            except Exception as e:
                artist_cols = [str(e)]
                
        return JsonResponse({
            "tables": tables,
            "migrations": migrations,
            "artist_columns": artist_cols
        })
    return render(request, "home.html")



from django.shortcuts import render


from django.shortcuts import render, redirect
from .models import Review


def reviews_page(request):

    if request.method == "POST":

        name = request.POST.get("name")
        email = request.POST.get("email")
        review = request.POST.get("review")
        rating = request.POST.get("rating")
        image = request.FILES.get("image")

        Review.objects.create(
            customer_name=name,
            customer_email=email,
            customer_review=review,
            rating=rating,
            review_image=image,
        )

        return redirect("reviews_page")

    reviews = Review.objects.all().order_by("-review_date")

    return render(request, "reviews.html", {"reviews": reviews})


def _require_staff_or_404(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        raise Http404("Page not found")


def admin_reports(request):
    from accounts.models import CustomUser, Employee, Customer
    from wallet.models import WalletTransaction
    from django.db.models import Q
    from employee.models import ServiceImage

    _require_staff_or_404(request)

    per_page = 20
    active_tab = request.GET.get("tab", "transactions")

    txn_qs = (
        WalletTransaction.objects.select_related("wallet__user")
        .exclude(wallet__user__is_staff=True)
        .exclude(wallet__user__is_superuser=True)
        .exclude(wallet__user__role__iexact="admin")
        .order_by("-created_at")
    )
    users_qs = (
        CustomUser.objects.exclude(is_staff=True)
        .exclude(is_superuser=True)
        .exclude(role__iexact="admin")
        .order_by("-date_joined")
    )
    employees_qs = (
        Employee.objects.select_related("user")
        .exclude(user__is_staff=True)
        .exclude(user__is_superuser=True)
        .exclude(user__role__iexact="admin")
        .order_by("-id")
    )
    customers_qs = (
        Customer.objects.select_related("user")
        .exclude(user__is_staff=True)
        .exclude(user__is_superuser=True)
        .exclude(user__role__iexact="admin")
        .order_by("-id")
    )
    pending_qs = ServiceImage.objects.filter(is_approved=False).order_by("-id")

    # Transactions filters
    txn_search = request.GET.get("txn_search", "").strip()
    txn_type = request.GET.get("txn_type", "").strip()
    if txn_search:
        txn_qs = txn_qs.filter(
            Q(transaction_id__icontains=txn_search)
            | Q(razorpay_payment_id__icontains=txn_search)
            | Q(wallet__user__email__icontains=txn_search)
        )
    if txn_type:
        txn_qs = txn_qs.filter(transaction_type__iexact=txn_type)

    # Users filters
    user_search = request.GET.get("user_search", "").strip()
    user_role = request.GET.get("user_role", "").strip()
    user_verified = request.GET.get("user_verified", "").strip()
    if user_search:
        users_qs = users_qs.filter(
            Q(email__icontains=user_search) | Q(full_name__icontains=user_search)
        )
    if user_role:
        users_qs = users_qs.filter(role__iexact=user_role)
    if user_verified in ("yes", "no"):
        users_qs = users_qs.filter(is_verified=(user_verified == "yes"))

    # Employees filters
    emp_search = request.GET.get("emp_search", "").strip()
    emp_status = request.GET.get("emp_status", "").strip()
    emp_verified = request.GET.get("emp_verified", "").strip()
    if emp_search:
        employees_qs = employees_qs.filter(
            Q(full_name__icontains=emp_search)
            | Q(email_address__icontains=emp_search)
            | Q(mobile__icontains=emp_search)
            | Q(user__email__icontains=emp_search)
        )
    if emp_status:
        employees_qs = employees_qs.filter(status__iexact=emp_status)
    if emp_verified in ("yes", "no"):
        employees_qs = employees_qs.filter(is_verified=(emp_verified == "yes"))

    # Customers filters
    cust_search = request.GET.get("cust_search", "").strip()
    cust_verified = request.GET.get("cust_verified", "").strip()
    if cust_search:
        customers_qs = customers_qs.filter(
            Q(customer_full_name__icontains=cust_search)
            | Q(email__icontains=cust_search)
            | Q(mobile__icontains=cust_search)
            | Q(user__email__icontains=cust_search)
        )
    if cust_verified in ("yes", "no"):
        customers_qs = customers_qs.filter(is_verified=(cust_verified == "yes"))

    # Pending approvals filters
    pending_search = request.GET.get("pending_search", "").strip()
    pending_media = request.GET.get("pending_media", "").strip()
    if pending_search:
        pending_qs = pending_qs.filter(
            Q(image_name__icontains=pending_search)
            | Q(type_of_art__icontains=pending_search)
            | Q(userupload_name__icontains=pending_search)
        )
    if pending_media == "photo":
        pending_qs = pending_qs.exclude(file_url__iregex=r"\.(mp4|mov|avi|webm|mkv|m4v)$")
    elif pending_media == "video":
        pending_qs = pending_qs.filter(file_url__iregex=r"\.(mp4|mov|avi|webm|mkv|m4v)$")

    txn_page = Paginator(txn_qs, per_page).get_page(request.GET.get("tpage"))
    users_page = Paginator(users_qs, per_page).get_page(request.GET.get("upage"))
    employees_page = Paginator(employees_qs, per_page).get_page(request.GET.get("epage"))
    customers_page = Paginator(customers_qs, per_page).get_page(request.GET.get("cpage"))
    pending_page = Paginator(pending_qs, per_page).get_page(request.GET.get("ppage"))

    return render(
        request,
        "admin_reports.html",
        {
            "active_tab": active_tab,
            "txn_page": txn_page,
            "users_page": users_page,
            "employees_page": employees_page,
            "customers_page": customers_page,
            "pending_page": pending_page,
            "txn_filter": {"search": txn_search, "type": txn_type},
            "user_filter": {
                "search": user_search,
                "role": user_role,
                "verified": user_verified,
            },
            "emp_filter": {
                "search": emp_search,
                "status": emp_status,
                "verified": emp_verified,
            },
            "cust_filter": {"search": cust_search, "verified": cust_verified},
            "pending_filter": {"search": pending_search, "media": pending_media},
        },
    )


def admin_reports_export(request, dataset):
    from accounts.models import CustomUser, Employee, Customer
    from wallet.models import WalletTransaction
    from openpyxl import Workbook
    from home.models import Booking, BookingOrder, Review, CustomProduct
    from employee.models import ServiceImage

    _require_staff_or_404(request)

    def safe_count(qs):
        try:
            return qs.count()
        except (ProgrammingError, OperationalError):
            return 0

    def safe_list(qs):
        try:
            return list(qs)
        except (ProgrammingError, OperationalError):
            return []

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{dataset.title()} Report"

    if dataset == "transactions":
        sheet.append(
            ["Transaction ID", "Type", "Amount", "User Email", "User Role", "Payment ID", "Created At", "Updated At"]
        )
        txn_rows = (
            WalletTransaction.objects.select_related("wallet__user")
            .exclude(wallet__user__is_staff=True)
            .exclude(wallet__user__is_superuser=True)
            .exclude(wallet__user__role__iexact="admin")
            .order_by("-created_at")
        )
        for txn in txn_rows:
            user = txn.wallet.user if txn.wallet else None
            sheet.append(
                [
                    txn.transaction_id,
                    txn.transaction_type,
                    float(txn.amount) if txn.amount is not None else 0,
                    getattr(user, "email", ""),
                    getattr(user, "role", ""),
                    txn.razorpay_payment_id or "",
                    txn.created_at.strftime("%Y-%m-%d %H:%M:%S") if txn.created_at else "",
                    txn.updated_at.strftime("%Y-%m-%d %H:%M:%S") if txn.updated_at else "",
                ]
            )
    elif dataset == "users":
        sheet.append(["ID", "Email", "Full Name", "Role", "Is Verified", "Is Staff", "Created At", "Updated At"])
        users_rows = (
            CustomUser.objects.exclude(is_staff=True)
            .exclude(is_superuser=True)
            .exclude(role__iexact="admin")
            .order_by("-date_joined")
        )
        for user in users_rows:
            sheet.append(
                [
                    user.id,
                    user.email,
                    user.full_name or "",
                    user.role or "",
                    user.is_verified,
                    user.is_staff,
                    user.date_joined.strftime("%Y-%m-%d %H:%M:%S") if user.date_joined else "",
                    user.last_login.strftime("%Y-%m-%d %H:%M:%S") if user.last_login else "",
                ]
            )
    elif dataset == "employees":
        sheet.title = "Employees Profile"
        sheet.append(
            [
                "ID",
                "Full Name",
                "Email",
                "Mobile",
                "Status",
                "Verified",
                "City",
                "State",
                "User Email",
                "Created At",
                "Updated At",
                "Total Bookings",
                "Total Wallet Transactions",
                "Total Uploads",
            ]
        )
        employees_rows = list(
            Employee.objects.select_related("user")
            .exclude(user__is_staff=True)
            .exclude(user__is_superuser=True)
            .exclude(user__role__iexact="admin")
            .order_by("-id")
        )
        employee_users = [e.user for e in employees_rows if e.user]
        booking_counts = {}
        wallet_counts = {}
        upload_counts = {}
        for emp in employees_rows:
            if emp.user:
                booking_counts[emp.user.id] = safe_count(Booking.objects.filter(assigned_employee=emp.user))
                wallet_counts[emp.user.id] = safe_count(WalletTransaction.objects.filter(wallet__user=emp.user))
                upload_counts[emp.user.id] = safe_count(ServiceImage.objects.filter(userupload_id=emp.user.id))

        for emp in employees_rows:
            user_id = emp.user.id if emp.user else None
            sheet.append(
                [
                    emp.id,
                    emp.full_name or "",
                    emp.email_address or "",
                    emp.mobile or "",
                    emp.status,
                    emp.is_verified,
                    emp.city or "",
                    emp.state or "",
                    emp.user.email if emp.user else "",
                    emp.user.date_joined.strftime("%Y-%m-%d %H:%M:%S") if emp.user and emp.user.date_joined else "",
                    emp.user.last_login.strftime("%Y-%m-%d %H:%M:%S") if emp.user and emp.user.last_login else "",
                    booking_counts.get(user_id, 0),
                    wallet_counts.get(user_id, 0),
                    upload_counts.get(user_id, 0),
                ]
            )

        bookings_sheet = workbook.create_sheet("Employees Bookings")
        bookings_sheet.append(
            [
                "Employee ID",
                "Employee Email",
                "Booking ID",
                "Service",
                "Customer Name",
                "Status",
                "Assignment Status",
                "Payment Amount",
                "Created At",
            ]
        )
        bookings_rows = Booking.objects.select_related("assigned_employee").filter(
            assigned_employee__in=employee_users
        ).order_by("-created_at")
        for b in bookings_rows:
            bookings_sheet.append(
                [
                    b.assigned_employee.id if b.assigned_employee else "",
                    b.assigned_employee.email if b.assigned_employee else "",
                    b.booking_id,
                    b.service_name,
                    b.customer_name,
                    b.status,
                    b.assignment_status,
                    float(b.payment_amount) if b.payment_amount is not None else 0,
                    b.created_at.strftime("%Y-%m-%d %H:%M:%S") if b.created_at else "",
                ]
            )

        wallet_sheet = workbook.create_sheet("Employees Wallet Txn")
        wallet_sheet.append(
            [
                "Employee ID",
                "Employee Email",
                "Transaction ID",
                "Type",
                "Amount",
                "Payment ID",
                "Created At",
            ]
        )
        wallet_rows = WalletTransaction.objects.select_related("wallet__user").filter(
            wallet__user__in=employee_users
        ).order_by("-created_at")
        for t in wallet_rows:
            user = t.wallet.user if t.wallet else None
            wallet_sheet.append(
                [
                    user.id if user else "",
                    user.email if user else "",
                    t.transaction_id,
                    t.transaction_type,
                    float(t.amount) if t.amount is not None else 0,
                    t.razorpay_payment_id or "",
                    t.created_at.strftime("%Y-%m-%d %H:%M:%S") if t.created_at else "",
                ]
            )

        upload_sheet = workbook.create_sheet("Employees Uploads")
        upload_sheet.append(
            [
                "Employee ID",
                "Uploader Name",
                "Image Name",
                "Type of Art",
                "Price",
                "Approved",
                "Verified Pic",
            ]
        )
        upload_rows = ServiceImage.objects.filter(
            userupload_id__in=[u.id for u in employee_users]
        ).order_by("-id")
        for up in upload_rows:
            upload_sheet.append(
                [
                    up.userupload_id,
                    up.userupload_name or "",
                    up.image_name or "",
                    up.type_of_art or "",
                    float(up.price) if up.price is not None else 0,
                    up.is_approved,
                    up.is_verified_pic,
                ]
            )
    elif dataset == "customers":
        sheet.title = "Customers Profile"
        sheet.append(
            [
                "ID",
                "Name",
                "Email",
                "Mobile",
                "Verified",
                "User Email",
                "Created At",
                "Updated At",
                "Total Bookings",
                "Total Wallet Transactions",
                "Total Orders",
                "Total Reviews",
                "Total Custom Products",
            ]
        )
        customers_rows = list(
            Customer.objects.select_related("user")
            .exclude(user__is_staff=True)
            .exclude(user__is_superuser=True)
            .exclude(user__role__iexact="admin")
            .order_by("-id")
        )
        customer_users = [c.user for c in customers_rows if c.user]
        for cust in customers_rows:
            bookings_count = safe_count(Booking.objects.filter(customer_user_id=cust.user.id)) if cust.user else 0
            wallet_count = safe_count(WalletTransaction.objects.filter(wallet__user=cust.user)) if cust.user else 0
            orders_count = safe_count(BookingOrder.objects.filter(user=cust.user)) if cust.user else 0
            reviews_count = safe_count(Review.objects.filter(customer_email__iexact=(cust.email or (cust.user.email if cust.user else ""))))
            custom_count = safe_count(CustomProduct.objects.filter(user=cust.user)) if cust.user else 0
            sheet.append(
                [
                    cust.id,
                    cust.customer_full_name,
                    cust.email or "",
                    cust.mobile or "",
                    cust.is_verified,
                    cust.user.email if cust.user else "",
                    cust.user.date_joined.strftime("%Y-%m-%d %H:%M:%S") if cust.user and cust.user.date_joined else "",
                    cust.user.last_login.strftime("%Y-%m-%d %H:%M:%S") if cust.user and cust.user.last_login else "",
                    bookings_count,
                    wallet_count,
                    orders_count,
                    reviews_count,
                    custom_count,
                ]
            )

        cust_bookings = workbook.create_sheet("Customers Bookings")
        cust_bookings.append(
            [
                "Customer User ID",
                "Customer Email",
                "Booking ID",
                "Service",
                "Status",
                "Payment Amount",
                "Total Amount",
                "Created At",
            ]
        )
        bookings_rows = Booking.objects.filter(customer_user_id__in=[u.id for u in customer_users]).order_by("-created_at")
        customer_email_map = {u.id: u.email for u in customer_users}
        for b in bookings_rows:
            cust_bookings.append(
                [
                    b.customer_user_id,
                    customer_email_map.get(b.customer_user_id, ""),
                    b.booking_id,
                    b.service_name,
                    b.status,
                    float(b.payment_amount) if b.payment_amount is not None else 0,
                    float(b.total_amount) if b.total_amount is not None else 0,
                    b.created_at.strftime("%Y-%m-%d %H:%M:%S") if b.created_at else "",
                ]
            )

        cust_wallet = workbook.create_sheet("Customers Wallet Txn")
        cust_wallet.append(
            ["Customer ID", "Customer Email", "Transaction ID", "Type", "Amount", "Payment ID", "Created At"]
        )
        wallet_rows = WalletTransaction.objects.select_related("wallet__user").filter(
            wallet__user__in=customer_users
        ).order_by("-created_at")
        for t in wallet_rows:
            user = t.wallet.user if t.wallet else None
            cust_wallet.append(
                [
                    user.id if user else "",
                    user.email if user else "",
                    t.transaction_id,
                    t.transaction_type,
                    float(t.amount) if t.amount is not None else 0,
                    t.razorpay_payment_id or "",
                    t.created_at.strftime("%Y-%m-%d %H:%M:%S") if t.created_at else "",
                ]
            )

        cust_orders = workbook.create_sheet("Customers Orders")
        cust_orders.append(
            ["Customer ID", "Customer Email", "Order Query ID", "Product", "Amount", "Status", "Created At"]
        )
        order_rows = BookingOrder.objects.select_related("user").filter(user__in=customer_users).order_by("-created_at")
        for o in order_rows:
            cust_orders.append(
                [
                    o.user.id if o.user else "",
                    o.user.email if o.user else "",
                    str(o.query_id),
                    o.product_name,
                    float(o.amount) if o.amount is not None else 0,
                    o.status,
                    o.created_at.strftime("%Y-%m-%d %H:%M:%S") if o.created_at else "",
                ]
            )

        cust_reviews = workbook.create_sheet("Customers Reviews")
        cust_reviews.append(
            ["Customer Name", "Customer Email", "Rating", "Review", "Review Date"]
        )
        review_emails = [c.email for c in customers_rows if c.email] + [u.email for u in customer_users]
        review_rows = safe_list(Review.objects.filter(customer_email__in=review_emails).order_by("-review_date"))
        for r in review_rows:
            cust_reviews.append(
                [
                    r.customer_name,
                    r.customer_email,
                    r.rating,
                    r.customer_review or "",
                    r.review_date.strftime("%Y-%m-%d %H:%M:%S") if r.review_date else "",
                ]
            )

        cust_custom = workbook.create_sheet("Customers Custom Products")
        cust_custom.append(
            ["Customer ID", "Customer Email", "Name", "Size", "Material", "Message", "Created At"]
        )
        custom_rows = CustomProduct.objects.select_related("user").filter(user__in=customer_users).order_by("-created_at")
        for cp in custom_rows:
            cust_custom.append(
                [
                    cp.user.id if cp.user else "",
                    cp.user.email if cp.user else "",
                    cp.name or "",
                    cp.size or "",
                    cp.material or "",
                    cp.message or "",
                    cp.created_at.strftime("%Y-%m-%d %H:%M:%S") if cp.created_at else "",
                ]
            )
    else:
        sheet.append(["Invalid dataset"])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{dataset}_report.xlsx"'
    workbook.save(response)
    return response


def admin_reports_detail(request, dataset, record_id):
    from accounts.models import Employee, Customer
    from wallet.models import WalletTransaction
    from home.models import Booking, BookingOrder, Review, CustomProduct
    from employee.models import ServiceImage

    _require_staff_or_404(request)

    def safe_list(qs):
        try:
            return list(qs)
        except (ProgrammingError, OperationalError):
            return []

    context = {"dataset": dataset}
    if dataset == "employees":
        emp = get_object_or_404(Employee.objects.select_related("user"), id=record_id)
        user = emp.user
        context["record"] = emp
        context["bookings"] = safe_list(Booking.objects.filter(assigned_employee=user).order_by("-created_at")) if user else []
        context["wallet_txns"] = safe_list(
            WalletTransaction.objects.select_related("wallet__user").filter(wallet__user=user).order_by("-created_at")
        ) if user else []
        context["uploads"] = safe_list(ServiceImage.objects.filter(userupload_id=user.id).order_by("-id")) if user else []
    elif dataset == "customers":
        cust = get_object_or_404(Customer.objects.select_related("user"), id=record_id)
        user = cust.user
        context["record"] = cust
        context["bookings"] = safe_list(Booking.objects.filter(customer_user_id=user.id).order_by("-created_at")) if user else []
        context["wallet_txns"] = safe_list(
            WalletTransaction.objects.select_related("wallet__user").filter(wallet__user=user).order_by("-created_at")
        ) if user else []
        context["orders"] = safe_list(BookingOrder.objects.filter(user=user).order_by("-created_at")) if user else []
        email_key = cust.email or (user.email if user else "")
        context["reviews"] = safe_list(Review.objects.filter(customer_email__iexact=email_key).order_by("-review_date")) if email_key else []
        context["custom_products"] = safe_list(CustomProduct.objects.filter(user=user).order_by("-created_at")) if user else []
    else:
        raise Http404("Invalid report dataset")

    return render(request, "home/admin_reports_detail.html", context)


@require_POST
def admin_reports_delete(request, dataset, record_id):
    from accounts.models import CustomUser, Employee, Customer

    _require_staff_or_404(request)

    next_url = request.POST.get("next") or reverse("admin_reports")
    if not str(next_url).startswith("/reports/"):
        next_url = reverse("admin_reports")
    if request.POST.get("confirm_text") != "DELETE":
        messages.error(request, "Delete blocked. You must type DELETE to confirm.")
        return redirect(next_url)

    if dataset == "users":
        obj = get_object_or_404(CustomUser, id=record_id)
        if obj.id == request.user.id:
            messages.error(request, "You cannot delete your own admin account.")
            return redirect(next_url)
        if obj.is_staff or obj.is_superuser:
            messages.error(request, "Staff/Admin users cannot be deleted from this page.")
            return redirect(next_url)
        obj.delete()
        messages.success(request, "User deleted successfully.")
    elif dataset == "employees":
        obj = get_object_or_404(Employee, id=record_id)
        obj.delete()
        messages.success(request, "Employee profile deleted successfully.")
    elif dataset == "customers":
        obj = get_object_or_404(Customer, id=record_id)
        obj.delete()
        messages.success(request, "Customer profile deleted successfully.")
    else:
        messages.error(request, "Invalid delete action.")

    return redirect(next_url)


from django.contrib.auth import logout

# from django.shortcuts import render, redirect


def logout_view(request):
    logout(request)
    return render(request, "accounts/login.html")


from django.shortcuts import render
from django.db.models import Q
from django.core.paginator import Paginator
from accounts.models import Employee


def artists(request):
    query = Employee.objects.all().order_by("-id")

    # Get filters
    artist_id = request.GET.get("artist_id")
    name = request.GET.get("name")
    pin_code = request.GET.get("pin_code")
    address = request.GET.get("address")
    work_type = request.GET.get("work_type")
    experience_years = request.GET.get("experience_years")

    # Filter by Artist ID (supports RAS5, ras12, 8 etc.)
    if artist_id:
        clean_id = (
            artist_id.upper().replace("RAS", "").strip()
        )  # ensures case-insensitive
        if clean_id.isdigit():
            query = query.filter(id=clean_id)

    # Apply filters with Q objects
    if name:
        query = query.filter(full_name__icontains=name)
    if pin_code:
        query = query.filter(pincode__icontains=pin_code)
    if address:
        query = query.filter(
            Q(village__icontains=address)
            | Q(city__icontains=address)
            | Q(state__icontains=address)
        )
    if work_type:
        query = query.filter(type_of_work__icontains=work_type)
    if experience_years:
        try:
            experience_years = int(experience_years)
            query = query.filter(experience__gte=experience_years)
        except ValueError:
            pass  # ignore invalid input

    paginator = Paginator(query, 8)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "artist.html",
        {
            "artists": page_obj.object_list,
            "page_obj": page_obj,
            "total_artists": paginator.count,
        },
    )


from .models import Review
import uuid

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from .models import Review


@csrf_exempt
@require_POST
def save_review(request):
    try:
        if not request.user.is_authenticated:
            return JsonResponse(
                {"success": False, "error": "User not authenticated"}, status=401
            )

        # Ensure logged-in user is a customer
        try:
            customer = Customer.objects.get(user=request.user)
        except Customer.DoesNotExist:
            return JsonResponse(
                {"success": False, "error": "Only customers can submit reviews"},
                status=403,
            )

        name = request.POST.get("name")
        email = request.POST.get("email")
        review_text = request.POST.get("customer_review")
        rating = request.POST.get("rating")
        review_image = request.FILES.get("image")

        if not (name and email and rating):
            return JsonResponse(
                {"success": False, "error": "Missing required fields"}, status=400
            )

        # ✅ Convert customer.id to string since customer_id is CharField
        review = Review.objects.create(
            customer_id=str(customer.id),  # Convert to string
            customer_name=name,
            customer_email=email,
            customer_review=review_text or "",
            rating=int(rating),  # Ensure integer
            review_image=review_image,
        )

        # Force save and verify
        review.save()

        # Debug print
        print(f"✅ Review saved successfully: {review.id} - {review.customer_name}")
        profile_pic_url = (
            customer.customer_photo.url if customer.customer_photo else None
        )

        return JsonResponse(
            {
                "success": True,
                "message": "Review submitted successfully!",
                "data": {
                    "customer_id": review.customer_id,
                    "name": review.customer_name,
                    "email": review.customer_email,
                    "customer_review": review.customer_review,
                    "rating": review.rating,
                    "profile_pic": profile_pic_url,
                    "image": review.review_image.url if review.review_image else None,
                    "created_at": review.review_date.strftime("%Y-%m-%d %H:%M"),
                },
            }
        )

    except ValueError as e:
        print(f"❌ ValueError: {e}")
        return JsonResponse(
            {"success": False, "error": f"Invalid rating value: {str(e)}"}, status=400
        )
    except Exception as e:
        print(f"❌ Exception: {e}")
        import traceback

        traceback.print_exc()
        return JsonResponse({"success": False, "error": str(e)}, status=500)


from .models import Booking

# def bookings(request):
#     status_filter = request.GET.get('status')  # ?status=pending, ?status=all
#     if status_filter and status_filter.lower() != "all":
#         bookings = Booking.objects.filter(status=status_filter)
#     else:
#         bookings = Booking.objects.all()

#     return render(request, 'bookings.html', {'bookings': bookings})
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from .models import Booking
from accounts.models import CustomUser  # Assuming CustomUser is in accounts app
from django.contrib.auth.decorators import user_passes_test
from django.db.models import Q


# def update_booking_status(request, booking_id):
#     if request.method == "POST":
#         new_status = request.POST.get("status")
#         booking = get_object_or_404(Booking, id=booking_id)
#         booking.status = new_status
#         booking.save()
#     return redirect(reverse("bookings"))

from django.http import JsonResponse
from .models import Booking
from accounts.models import Customer
from datetime import datetime
import uuid
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required

# @login_required
# @csrf_exempt
# def save_booking(request):
#     if request.method == 'POST':
#         try:
#             print("Logged-in user:", request.user, request.user.id)

#             # Get the Customer record
#             customer_qs = request.user.customers.all()
#             if not customer_qs.exists():
#                 return JsonResponse({'success': False, 'message': 'This option is only available for customers.'})

#             customer = customer_qs.first()  # the single Customer object

#             booking_id = str(uuid.uuid4())[:8]

#             appointment_date = None
#             appointment_date_str = request.POST.get('appointment_date')
#             if appointment_date_str:
#                 appointment_date = datetime.strptime(appointment_date_str, "%d-%m-%Y").date()

#             booking = Booking.objects.create(
#                 booking_id=booking_id,
#                 customer_name=customer.customer_full_name,
#                 customer_user_id=customer.id,
#                 service_name=request.POST.get('service_name'),
#                 contact_number=request.POST.get('contact_number'),
#                 email=request.POST.get('email'),
#                 address=request.POST.get('address'),
#                 pin_code=request.POST.get('pin_code'),
#                 state=request.POST.get('state'),
#                 city=request.POST.get('city'),
#                 total_walls=int(request.POST.get('total_walls', 0)),
#                 width=float(request.POST.get('width', 0)),
#                 height=float(request.POST.get('height', 0)),
#                 total_sqft=float(request.POST.get('total_sqft', 0)),
#                 appointment_date=appointment_date,
#                 payment_option="pending",
#                 payment_amount=0.00,
#                 is_paid=False
#             )

#             return JsonResponse({'success': True, 'message': 'Booking saved successfully!'})

#         except Exception as e:
#             return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})


from django.shortcuts import render
from django.contrib.auth.decorators import login_required

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from accounts.models import Customer, Employee  # assuming you have Employee model
from decimal import Decimal
import time
from decimal import Decimal
from wallet.models import Wallet, WalletTransaction
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required


@login_required
def my_uploads_view(request):
    if request.user.role != "employee":
        messages.error(request, "This page is only available for employee accounts.")
        return redirect("home")

    uploads = ServiceImage.objects.filter(userupload_id=request.user.id).order_by("-id")
    video_exts = (".mp4", ".mov", ".avi", ".webm", ".mkv", ".m4v")

    def _media_url(item):
        if item.file_url:
            return str(item.file_url).lower()
        try:
            return str(item.image.url).lower()
        except Exception:
            return str(item.image).lower()

    photo_uploads = []
    video_uploads = []
    for item in uploads:
        media_url = _media_url(item)
        if media_url.endswith(video_exts):
            video_uploads.append(item)
        else:
            photo_uploads.append(item)

    context = {
        "uploads": uploads,
        "photo_uploads": photo_uploads,
        "video_uploads": video_uploads,
    }
    return render(request, "my_uploads.html", context)


@login_required
def edit_profile_view(request):
    # List of all 29 Indian states
    INDIAN_STATES = [
        "Andhra Pradesh",
        "Arunachal Pradesh",
        "Assam",
        "Bihar",
        "Chhattisgarh",
        "Goa",
        "Gujarat",
        "Haryana",
        "Himachal Pradesh",
        "Jharkhand",
        "Karnataka",
        "Kerala",
        "Madhya Pradesh",
        "Maharashtra",
        "Manipur",
        "Meghalaya",
        "Mizoram",
        "Nagaland",
        "Odisha",
        "Punjab",
        "Rajasthan",
        "Sikkim",
        "Tamil Nadu",
        "Telangana",
        "Tripura",
        "Uttar Pradesh",
        "Uttarakhand",
        "West Bengal",
        "Delhi",
    ]

    user = request.user
    if user.role == "employee":
        employee = get_object_or_404(Employee, user=user)

        if request.method == "POST":

            prev_status = employee.status
            employee.fathers_name = request.POST.get("fathers_name") or None
            employee.dob = request.POST.get("dob")
            employee.gender = request.POST.get("gender")
            employee.house_no = request.POST.get("house_no")
            employee.village = request.POST.get("village")
            employee.city = request.POST.get("city")
            employee.state = request.POST.get("state")
            raw_pincode = (request.POST.get("pincode") or "").strip()
            if raw_pincode and not raw_pincode.isdigit():
                return JsonResponse(
                    {"success": False, "message": "Pincode must contain only numbers."}
                )
            employee.pincode = raw_pincode or None
            employee.aadhar_card_no = request.POST.get("aadhar_card_no")
            employee.experience = request.POST.get("experience")

            # Get multiple selected locations and join them
            selected_locations = request.POST.getlist(
                "preferred_work_location"
            ) or request.POST.getlist("preferred_location")
            print("LOCATIONS:", selected_locations)
            employee.preferred_work_location = ", ".join(selected_locations)

            employee.bank_account_holder_name = (
                request.POST.get("bank_account_holder_name") or None
            )

            employee.account_no = request.POST.get("account_no") or None
            raw_ifsc = (request.POST.get("ifsc_code") or "").strip()
            employee.ifsc_code = raw_ifsc or None

            # NEW: Handle the three new optional fields
            employee.full_address = request.POST.get("full_address") or None
            employee.working_range = request.POST.get("working_range")
            employee.belong_to_org = request.POST.get("belong_to_org") == "on"
            employee.pan_card = request.POST.get("pan_card")
            employee.gst_no = request.POST.get("gst_no")
            employee.organization_name = request.POST.get("organization_name")
            employee.organization_type = request.POST.get("organization_type") or None

            employee.role = "employee"

            # Check if ready to take orders
            is_ready = request.POST.get("ready_to_take_orders") == "on"

            # File fields
            if "passport_photo" in request.FILES:
                employee.passport_photo = request.FILES["passport_photo"]
            if "aadhar_card_image_front" in request.FILES:
                employee.aadhar_card_image_front = request.FILES[
                    "aadhar_card_image_front"
                ]
            if "aadhar_card_image_back" in request.FILES:
                employee.aadhar_card_image_back = request.FILES[
                    "aadhar_card_image_back"
                ]

            # Multi checkbox values (list)
            employee.type_of_work = request.POST.getlist("type_of_work")

            # Handle status change and wallet deduction
            if not prev_status and is_ready:
                wallet = Wallet.objects.filter(user=user).first()

                if wallet and wallet.balance >= Decimal("50.00"):
                    wallet.balance -= Decimal("50.00")
                    wallet.save()

                    WalletTransaction.objects.create(
                        wallet=wallet,
                        transaction_type="DEBIT",
                        amount=Decimal("50.00"),
                        razorpay_payment_id=f"ACTIVATION_FEE_{user.id}_{int(time.time())}",
                    )

                    employee.status = True
                    employee.save()

                    print("Deducted ₹50 for activation fee")
                    return JsonResponse(
                        {
                            "success": True,
                            "message": "Profile updated successfully! ₹50 activation fee deducted from your wallet.",
                        }
                    )
                else:
                    employee.status = False
                    employee.save()
                    return JsonResponse(
                        {
                            "success": False,
                            "message": "Insufficient wallet balance. Please add ₹50 to enable 'Ready to Take Orders' feature.",
                        }
                    )
            else:
                employee.full_name = request.POST.get("full_name")
                employee.status = is_ready
                employee.save()
                print("Profile saved")
                return JsonResponse(
                    {
                        "success": True,
                        "message": "Profile updated successfully!",
                        "name": employee.full_name,
                    }
                )

        # For GET requests - prepare context
        stored_locations = employee.preferred_work_location or ""
        selected_locations = (
            [loc.strip() for loc in stored_locations.split(",")]
            if stored_locations
            else []
        )

        context = {
            "name": employee.full_name,
            "email_address": employee.email_address,
            "contact": employee.mobile,
            "passport_photo": (
                employee.passport_photo.url if employee.passport_photo else None
            ),
            "fathers_name": employee.fathers_name,
            "dob": employee.dob,
            "gender": employee.gender,
            "house_no": employee.house_no,
            "village": employee.village,
            "city": employee.city,
            "state": employee.state,
            "pincode": employee.pincode,
            "aadhar_card_no": employee.aadhar_card_no,
            "aadhar_card_image_front": (
                employee.aadhar_card_image_front.url
                if employee.aadhar_card_image_front
                else None
            ),
            "aadhar_card_image_back": (
                employee.aadhar_card_image_back.url
                if employee.aadhar_card_image_back
                else None
            ),
            "experience": employee.experience,
            "type_of_work": employee.type_of_work or [],
            "preferred_work_location": employee.preferred_work_location,
            "INDIAN_STATES": INDIAN_STATES,
            "selected_locations": selected_locations,
            "bank_account_holder_name": employee.bank_account_holder_name,
            "account_no": employee.account_no,
            "ifsc_code": employee.ifsc_code,
            # NEW: Add the three new fields to context
            "pan_card": employee.pan_card,
            "gst_no": employee.gst_no,
            "organization_name": employee.organization_name,
            "organization_type": employee.organization_type,
            "ready_to_take_orders": employee.status,
            "full_address": employee.full_address,
            "working_range": employee.working_range,
            "belong_to_org": employee.belong_to_org,
        }
        return render(request, "edit_profile.html", context)

    elif user.role == "customer":
        customer = get_object_or_404(Customer, user=user)

        if request.method == "POST":
            customer.customer_full_name = request.POST.get("name")
            customer.email = request.POST.get("email") or None
            customer.mobile = request.POST.get("contact")

            # Handle profile picture
            if "customer_photo" in request.FILES:
                customer.customer_photo = request.FILES["customer_photo"]

            customer.save()
            return JsonResponse(
                {"success": True, "message": "Profile updated successfully!"}
            )

        context = {
            "name": customer.customer_full_name,
            "email": customer.email,
            "contact": customer.mobile,
            "profile_pic": (
                customer.customer_photo.url if customer.customer_photo else None
            ),
        }
        return render(request, "edit_customers_profile.html", context)

    return render(request, "edit_profile.html", {"user": user})


def shop(request):
    # Example products
    products = [
        {
            "id": 1,
            "name": "Acrylic Paint Set",
            "price": 599,
            "image": "gallery/paint_set.jpg",
        },
        {"id": 2, "name": "Canvas Board", "price": 299, "image": "gallery/canvas.jpg"},
        {"id": 3, "name": "Brush Kit", "price": 399, "image": "gallery/brush_kit.jpg"},
        {"id": 4, "name": "Oil Pastels", "price": 199, "image": "gallery/pastels.jpg"},
    ]
    return render(request, "shop.html", {"products": products})


from django.shortcuts import render
from .models import Booking


def my_orders(request):
    if not request.user.is_authenticated or request.user.role != "customer":
        return render(request, "not_allowed.html")

    # Filter bookings based on the logged-in user's ID
    bookings = Booking.objects.filter(customer_user_id=request.user.id).order_by(
        "-created_at"
    )

    return render(request, "my_orders.html", {"bookings": bookings})


# views.py

from django.shortcuts import render
from .models import Review
from django.db.models import F
from .models import Review
from accounts.models import Customer


def home_view(request):
    reviews = Review.objects.all().order_by("-review_date")[:3]
    reviews_list = list(reviews)

    fake_reviews = [
        {
            "customer_name": "P. Reddy",
            "customer_review": "A true artist! The mural they painted for my cafe is a masterpiece.",
            "rating": 5,
            "review_image": None,
        },
        {
            "customer_name": "S. Kumar",
            "customer_review": "Excellent work and very easy to communicate with. Will definitely use their service again.",
            "rating": 5,
            "review_image": None,
        },
        {
            "customer_name": "M. Gupta",
            "customer_review": "The 3D art on my wall looks so real. It completely changed the feel of the room.",
            "rating": 5,
            "review_image": None,
        },
    ]
    # keep your same fake reviews


def home_view(request):

    reviews_list = Review.objects.all().order_by("-id")

    # Fake reviews (अगर database में कम हों)
    fake_reviews = [
        {
            "customer_name": "Rohit Sharma",
            "rating": 5,
            "review": "Amazing wall painting work!",
            "review_image": None,
        },
        {
            "customer_name": "Pawan Mehta",
            "rating": 4,
            "review": "Very creative design and good finishing.",
            "review_image": None,
        },
        {
            "customer_name": "Neha Kapoor",
            "rating": 5,
            "review": "Professional artists and beautiful artwork.",
            "review_image": None,
        },
    ]

    all_reviews = list(reviews_list)

    if len(all_reviews) < 3:
        all_reviews = all_reviews + fake_reviews

    enriched_reviews = []

    for r in all_reviews:

        # Fake reviews (dictionary)
        if isinstance(r, dict):

            rating = r["rating"]
            r["full_stars"] = range(rating)
            r["empty_stars"] = range(5 - rating)
            r["customer_photo"] = None

            enriched_reviews.append(r)

        # Database reviews (Django model)
        else:

            rating = r.rating

            setattr(r, "full_stars", range(rating))
            setattr(r, "empty_stars", range(5 - rating))

            enriched_reviews.append(r)

    context = {"reviews": enriched_reviews, "premium_services": _premium_services_queryset()}

    return render(request, "home.html", context)


# views.py - Add these functions to your existing views.py

import razorpay
import time
import json
from django.conf import settings
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

# Initialize Razorpay client
razorpay_client = razorpay.Client(
    auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET)
)


@require_http_methods(["POST"])
def create_razorpay_order(request):
    try:
        # Get amount from request (convert to paise - multiply by 100)
        amount_str = request.POST.get("amount", "0")
        amount = int(float(amount_str) * 100)  # Convert to paise

        # Create unique receipt ID
        receipt_id = f"booking_{int(time.time())}"

        # Create Razorpay order
        order_data = {
            "amount": amount,
            "currency": "INR",
            "receipt": receipt_id,
            "payment_capture": 1,  # Auto capture payment
        }

        order = razorpay_client.order.create(data=order_data)

        return JsonResponse(
            {
                "success": True,
                "order_id": order["id"],
                "amount": order["amount"],
                "currency": order["currency"],
                "key_id": settings.RAZORPAY_KEY_ID,
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@csrf_exempt
@require_http_methods(["POST"])
def verify_razorpay_payment(request):
    try:
        data = json.loads(request.body)

        razorpay_order_id = data.get("razorpay_order_id")
        razorpay_payment_id = data.get("razorpay_payment_id")
        razorpay_signature = data.get("razorpay_signature")
        product_name = data.get("product_name")
        amount = data.get("amount")

        params_dict = {
            "razorpay_order_id": razorpay_order_id,
            "razorpay_payment_id": razorpay_payment_id,
            "razorpay_signature": razorpay_signature,
        }

        # Verify signature
        razorpay_client.utility.verify_payment_signature(params_dict)

        # 📝 Save booking/order
        from .models import BookingOrder  # or your booking model

        BookingOrder.objects.create(
            user=request.user if request.user.is_authenticated else None,
            product_name=product_name,
            amount=amount,
            razorpay_payment_id=razorpay_payment_id,
            razorpay_order_id=razorpay_order_id,
            status="PAID",
        )

        return JsonResponse(
            {"success": True, "message": "Payment verified successfully"}
        )

    except razorpay.errors.SignatureVerificationError:
        return JsonResponse(
            {"success": False, "error": "Payment signature verification failed"}
        )


# Add this to your views.py or update your existing save_bookings view
from django.core.mail import send_mail
from django.conf import settings

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
from .models import Booking


@csrf_exempt
@require_http_methods(["POST"])
def save_booking(request):
    try:
        if not request.user.is_authenticated:
            return JsonResponse(
                {"success": False, "message": "You must be logged in to book a service"}
            )
        if request.user.role != "customer":
            return JsonResponse(
                {
                    "success": False,
                    "message": "Only customer accounts can create bookings.",
                }
            )
        print("POST DATA:", request.POST)

        # --- Standard Fields ---
        service_name = request.POST.get("service_name")
        customer_name = request.POST.get("customer_name")
        print("NAME:", customer_name)
        # 👈 YAHAN ADD KARO
        contact_number = request.POST.get("contact_number")
        email = request.POST.get("email")
        address = request.POST.get("address")
        if not contact_number or not address:
            return JsonResponse(
                {"success": False, "message": "Contact & Address required"}
            )
        pin_code = request.POST.get("pin_code")
        state = request.POST.get("state")
        city = request.POST.get("city")
        total_walls = request.POST.get("total_walls") or "0"
        total_floor = request.POST.get("total_floor") or "0"
        width = request.POST.get("width")
        height = request.POST.get("height")
        width = float(width) if width else None
        height = float(height) if height else None
        total_sqft = request.POST.get("total_sqft") or "0"
        appointment_date = request.POST.get("appointment_date")

        # *** CRITICAL FIX: Use the calculated amount from the front-end ***
        total_amount = request.POST.get("total_amount")
        print(total_amount, "total_amount_+++++++++++++++++++++")
        # --- Design Fields (from form or null) ---
        selected_designs_raw = request.POST.get("selected_designs", "")
        custom_design_file = request.FILES.get(
            "custom_design"
        )  # This handles uploaded file

        try:
            selected_designs = json.loads(selected_designs_raw) if selected_designs_raw else []
        except json.JSONDecodeError:
            return JsonResponse(
                {"success": False, "message": "Invalid design selection data."}
            )

        if not isinstance(selected_designs, list) or len(selected_designs) == 0:
            return JsonResponse(
                {
                    "success": False,
                    "message": "Please select at least one design before booking.",
                }
            )

        try:
            total_walls = int(total_walls)
            total_floor = int(total_floor)
        except (TypeError, ValueError):
            return JsonResponse(
                {
                    "success": False,
                    "message": "Total Wall and Total Floor must be valid numbers.",
                }
            )

        if total_walls < 0 or total_floor < 0:
            return JsonResponse(
                {
                    "success": False,
                    "message": "Total Wall and Total Floor cannot be negative.",
                }
            )

        if total_walls == 0 and total_floor == 0:
            return JsonResponse(
                {
                    "success": False,
                    "message": "Please enter at least one Total Wall or Total Floor.",
                }
            )

        required_fields = [
            service_name,
            contact_number,
            address,
            pin_code,
            state,
            city,
            width,
            height,
            appointment_date,
        ]
        if not all(required_fields):
            # Check for total_sqft > 0 instead of just existence
            if not total_sqft or float(total_sqft) <= 0:
                return JsonResponse(
                    {
                        "success": False,
                        "message": "Please enter valid width/height to calculate square footage.",
                    }
                )
            return JsonResponse(
                {"success": False, "message": "Please fill all required fields"}
            )

        if appointment_date:
            try:
                # Assuming the JS sends 'DD-MM-YYYY'
                appointment_date = datetime.strptime(
                    appointment_date, "%d-%m-%Y"
                ).date()
            except ValueError:
                # Check if format is YYYY-MM-DD (standard date input fallback)
                try:
                    appointment_date = datetime.strptime(
                        appointment_date, "%Y-%m-%d"
                    ).date()
                except ValueError:
                    return JsonResponse(
                        {
                            "success": False,
                            "message": "Invalid date format, use DD-MM-YYYY or YYYY-MM-DD",
                        }
                    )

        next_id = Booking.objects.count() + 1
        booking_id = f"RCC{next_id}"
        print("getting to it ")

        # --- Finalize Design Fields for Model ---
        design_name_to_save = ", ".join(
            [str(item.get("name", "")).strip() for item in selected_designs if item.get("name")]
        ) or None

        price_to_save = sum(
            float(item.get("price", 0) or 0) for item in selected_designs
        )

        # Determine type_of_art_booked
        if design_name_to_save:
            art_type = "Selected Design"
        else:
            art_type = "Standard Service"
            # If standard, the price/rate is the hardcoded base rate from your JS, which you may want to save here.
            # For now, we'll keep price_to_save as None/0 if no design selected/uploaded.
            # If you want to save the rate from SERVICE_BASE_RATES, you'd need to send it from JS or look it up here.

        booking = Booking.objects.create(
            customer_name=(
                request.user.full_name if request.user.full_name else request.user.email
            ),
            customer_user_id=request.user.id,
            booking_id=booking_id,
            service_name=service_name,
            contact_number=contact_number,
            email=email,
            address=address,
            pin_code=pin_code,
            state=state,
            city=city,
            total_walls=total_walls,
            width=width,
            height=height,
            total_sqft=total_sqft,
            appointment_date=appointment_date,
            # --- SAVING DESIGN AND PRICE ---
            design_names=design_name_to_save,
            type_of_art_booked=art_type,
            price_of_design=price_to_save,  # The per-sqft rate or design fixed price
            customer_design=custom_design_file,  # Saves the uploaded file or None
            # --- FINAL AMOUNT FIX ---
            total_amount=total_amount,
        )

        # --- SEND BOOKING CONFIRMATION EMAIL ---
        try:
            subject = f"Booking Confirmation - {booking.booking_id}"

            message = f"""
        Hello {booking.customer_name},

        Thank you for booking with RColorCraft! 🎨  
        Your booking has been successfully received.

        Here are your booking details:

        ----------------------------------------
        🆔 Booking ID: {booking.booking_id}
        🛠 Service: {booking.service_name}

        📅 Appointment Date: {appointment_date.strftime('%d-%m-%Y')}
        🏠 Address: {booking.address}, {booking.city}, {booking.state} - {booking.pin_code}

        🧱 Total Walls: {booking.total_walls}
        🪵 Total Floors: {total_floor}
        📐 Width x Height: {booking.width} ft x {booking.height} ft
        📏 Total Sq Ft: {booking.total_sqft}

        💰 Total Amount: ₹{booking.total_amount}

        🎨 Art Type: {booking.type_of_art_booked}
        """

            # Add design info only if applicable
            if booking.design_names:
                message += f"🖼 Selected Design: {booking.design_names}\n"
                message += f"💵 Design Rate: ₹{booking.price_of_design} per sqft\n"

            if booking.customer_design:
                message += f"📁 Custom Design Uploaded: Yes\n"

            message += "\n----------------------------------------\n"
            message += "Our team will contact you within 24 hours.\n"
            message += "Thank you for choosing RColorCraft! 😊"

            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [email],  # customer email
                fail_silently=False,
            )

            print("BOOKING SAVED:", booking)
        except Exception as email_error:
            print("EMAIL ERROR:", email_error)

        return JsonResponse(
            {
                "success": True,
                "name": customer_name,
                "booking_id": booking.booking_id,
                "service": booking.service_name,
                "amount": booking.total_amount,
            }
        )

    except Exception as e:
        import traceback

        print(traceback.format_exc())
        return JsonResponse({"success": False, "message": str(e)})


from django.utils.decorators import method_decorator
from .models import CustomProduct


@csrf_exempt
@login_required
def save_custom_product(request):
    if request.method == "POST":
        try:
            CustomProduct.objects.create(
                user=request.user,
                name=request.POST.get("product_name"),
                size=request.POST.get("size"),
                material=request.POST.get("material"),
                other_material=request.POST.get("other_material"),
                message=request.POST.get("message"),
            )
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Invalid request"})


def is_admin(user):
    # Assuming staff/superuser status denotes an admin for booking management
    return user.is_authenticated and user.is_staff


def is_employee(user):
    # Assuming 'role' = 'employee' is set in CustomUser for workers
    return user.is_authenticated and user.role == "employee"


# --- Admin/Manager Views ---


@user_passes_test(is_admin)
def bookings(request):
    """Admin view to list all bookings and employees."""
    status_filter = request.GET.get("status")

    if status_filter and status_filter.lower() != "all":
        # Also filter by assignment_status if you want to group them
        bookings = Booking.objects.filter(
            Q(status=status_filter) | Q(assignment_status=status_filter)
        ).order_by("-created_at")
    else:
        bookings = Booking.objects.all().order_by("-created_at")

    # Fetch users who can be assigned (employees)
    employees = CustomUser.objects.filter(role="employee").order_by("full_name")

    context = {
        "bookings": bookings,
        "employees": employees,  # Passed for the assignment dropdown
    }
    return render(request, "bookings.html", context)


@user_passes_test(is_admin)
def update_booking_status(request, booking_id):
    """Admin view to change the main booking status (pending, in_process, etc.)."""
    if request.method == "POST":
        new_status = request.POST.get("status")
        booking = get_object_or_404(Booking, id=booking_id)
        booking.status = new_status
        booking.save()
    return redirect(reverse("bookings"))


# @user_passes_test(is_admin)
# def assign_booking(request, booking_id):
#     """Admin view to assign a booking to a specific employee."""
#     if request.method == "POST":
#         employee_id = request.POST.get("employee_id")
#         booking = get_object_or_404(Booking, id=booking_id)
#         employee = get_object_or_404(CustomUser, id=employee_id)

#         # Assign the employee and set status to 'assigned' (awaiting response)
#         booking.assigned_employee = employee
#         booking.assignment_status = 'assigned'
#         booking.save()

#     return redirect(reverse("bookings"))
from django.core.mail import send_mail
from django.conf import settings


@user_passes_test(is_admin)
def assign_booking(request, booking_id):
    """Admin view to assign a booking to a specific employee."""
    if request.method == "POST":
        employee_id = request.POST.get("employee_id")
        booking = get_object_or_404(Booking, id=booking_id)
        employee = get_object_or_404(CustomUser, id=employee_id)

        # Assign the employee
        booking.assigned_employee = employee
        booking.assignment_status = "assigned"
        booking.save()

        # 🔍 Fetch employee profile to get email
        employee_profile = Employee.objects.filter(user=employee).first()
        employee_email = employee_profile.email_address if employee_profile else None

        # 🔥 Send assignment email
        if employee_email:
            service = booking.service_name
            booking_date = booking.appointment_date
            city = booking.city
            state = booking.state

            msg = f"""
Dear {employee.full_name or employee.email},

You have received a new booking request.

🔹 Service: {service}
🔹 Date: {booking_date}
🔹 City: {city}
🔹 State: {state}

⚠️ Customer privacy protection:
- Customer phone number: ❌ Hidden
- Customer email: ❌ Hidden
- Customer full address: ❌ Hidden

For more information and full details,
👉 Please login to your account and accept the order.

Regards,
RColorcraft Bookings Team
"""

            send_mail(
                subject="📢 New Booking Assigned — Action Required",
                message=msg,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[employee_email],
                fail_silently=False,
            )

    return redirect(reverse("bookings"))


# --- Employee Views ---


@user_passes_test(is_employee)
def employee_bookings(request):
    """Employee view to see all assigned bookings."""
    # Only show bookings assigned to the currently logged-in user
    bookings = Booking.objects.filter(assigned_employee=request.user).order_by(
        "-created_at"
    )

    return render(request, "employee_bookings.html", {"bookings": bookings})


@user_passes_test(is_employee)
def handle_assignment_response(request, booking_id, action):
    """Employee view to accept or decline an assignment."""
    booking = get_object_or_404(Booking, id=booking_id)
    employee = request.user  # the logged-in user

    # Security check
    if booking.assigned_employee != employee or booking.assignment_status != "assigned":
        messages.error(request, "Invalid action or unauthorized access.")
        return redirect(reverse("employee_bookings"))

    # Accept booking
    if action == "accept":
        # Deduct ₹60 from wallet
        wallet = Wallet.objects.filter(user=employee).first()

        if wallet and wallet.balance >= Decimal("60.00"):
            wallet.balance -= Decimal("60.00")
            wallet.save()

            # Log transaction
            WalletTransaction.objects.create(
                wallet=wallet,
                transaction_type="DEBIT",
                amount=Decimal("60.00"),
                # description=f"Assignment fee for booking {booking.booking_id}",
                razorpay_payment_id=None,
            )

            booking.assignment_status = "accepted"
            booking.save()
            messages.success(
                request,
                f"You accepted the assignment. ₹60 has been deducted from your wallet.",
            )
        else:
            messages.error(
                request,
                "Insufficient wallet balance. You need at least ₹60 to accept this booking.",
            )
            return redirect(reverse("employee_bookings"))

    # Decline booking
    elif action == "decline":
        booking.assignment_status = "declined"
        booking.assigned_employee = None
        booking.save()
        messages.info(request, "You have declined the assignment.")

    return redirect(reverse("employee_bookings"))


from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json
from employee.models import ServiceImage


@csrf_exempt
def update_service_price(request):
    if request.method == "POST":
        if not request.user.is_staff:
            return JsonResponse(
                {"success": False, "message": "Permission denied."}, status=403
            )
        try:
            data = json.loads(request.body)
            image_id = data.get("id")
            new_price = data.get("price")

            img = ServiceImage.objects.get(id=image_id)
            img.price = new_price
            img.save()

            return JsonResponse({"success": True})
        except ServiceImage.DoesNotExist:
            return JsonResponse({"success": False, "message": "Image not found."})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)})
    return JsonResponse({"success": False, "message": "Invalid request method."})


from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from accounts.models import Employee


@login_required
def toggle_block_artist(request, artist_id):
    print("CLICKED ID:", artist_id)

    artist = get_object_or_404(Employee, id=artist_id)

    print("BEFORE:", artist.block_status)

    artist.block_status = not artist.block_status
    artist.save()

    print("AFTER:", artist.block_status)

    return redirect(request.META.get("HTTP_REFERER", "/artists/"))


from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Booking


@login_required
def toggle_customer_status(request, booking_id):
    booking = get_object_or_404(
        Booking, id=booking_id, customer_user_id=request.user.id
    )

    if request.method == "POST":
        booking.customer_status = not booking.customer_status
        booking.save(update_fields=["customer_status"])

    return redirect("my_orders")


@login_required
def toggle_artist_status(request, booking_id):
    booking = get_object_or_404(
        Booking,
        id=booking_id,
        assigned_employee=request.user,
        assignment_status="accepted",
    )

    if request.method == "POST":
        booking.artist_status = not booking.artist_status
        booking.save(update_fields=["artist_status"])

    return redirect("employee_assignments")


from django.shortcuts import render, redirect
from .models import Consultation


def home(request):
    if request.GET.get("run_repair") == "1":
        from django.db import connection
        from django.http import JsonResponse
        
        messages = []
        sql_commands = [
            # 1. Rename review to home_review if it exists
            "ALTER TABLE IF EXISTS review RENAME TO home_review;",
            
            # 2. Create home_consultation table
            """
            CREATE TABLE IF NOT EXISTS home_consultation (
                id bigserial PRIMARY KEY,
                name varchar(100) NOT NULL,
                email varchar(254) NOT NULL,
                phone varchar(15) NOT NULL,
                message text NOT NULL
            );
            """,
            
            # 3. Create home_customer table
            """
            CREATE TABLE IF NOT EXISTS home_customer (
                id bigserial PRIMARY KEY,
                user_id integer NOT NULL UNIQUE REFERENCES accounts_customuser(id) DEFERRABLE INITIALLY DEFERRED
            );
            """,
            
            # 4. Create home_artist table
            """
            CREATE TABLE IF NOT EXISTS home_artist (
                id bigserial PRIMARY KEY,
                user_id integer NOT NULL UNIQUE REFERENCES accounts_customuser(id) DEFERRABLE INITIALLY DEFERRED,
                is_active boolean NOT NULL DEFAULT FALSE,
                kyc_status varchar(20) NOT NULL DEFAULT 'pending'
            );
            """,
            
            # 5. Create home_assignment table
            """
            CREATE TABLE IF NOT EXISTS home_assignment (
                id bigserial PRIMARY KEY,
                customer_name varchar(100) NOT NULL,
                city varchar(50) NOT NULL,
                pincode varchar(10) NOT NULL,
                price integer NOT NULL,
                status varchar(20) NOT NULL DEFAULT 'pending',
                is_completed boolean NOT NULL DEFAULT FALSE
            );
            """,
            
            # 6. Re-insert migration history records to sync Django state
            """
            INSERT INTO django_migrations (app, name, applied) VALUES
            ('home', '0011_alter_review_id_alter_review_table', NOW()),
            ('home', '0012_consultation', NOW()),
            ('home', '0014_artist_customer', NOW()),
            ('home', '0015_assignment', NOW()),
            ('home', '0016_artist_kyc_status_artist_status', NOW()),
            ('home', '0018_rename_status_artist_is_active_and_more', NOW())
            ON CONFLICT (app, name) DO NOTHING;
            """
        ]
        
        with connection.cursor() as cursor:
            for sql in sql_commands:
                try:
                    cursor.execute(sql)
                    messages.append(f"Successfully executed SQL block.")
                except Exception as e:
                    messages.append(f"SQL Error: {str(e)}")
            
        return JsonResponse({
            "status": "success",
            "messages": messages
        })


    if request.GET.get("debug_db") == "1":
        from django.db import connection
        from django.http import JsonResponse
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT table_name 
                FROM information_schema.tables 
                WHERE table_schema='public'
            """)
            tables = [row[0] for row in cursor.fetchall()]
            
            cursor.execute("SELECT app, name, applied FROM django_migrations ORDER BY applied DESC")
            migrations = [f"{row[0]}: {row[1]} ({row[2]})" for row in cursor.fetchall()]
            
            artist_cols = []
            try:
                cursor.execute("""
                    SELECT column_name, data_type 
                    FROM information_schema.columns 
                    WHERE table_name = 'home_artist'
                """)
                artist_cols = [f"{row[0]} ({row[1]})" for row in cursor.fetchall()]
            except Exception as e:
                artist_cols = [str(e)]
                
        return JsonResponse({
            "tables": tables,
            "migrations": migrations,
            "artist_columns": artist_cols
        })

    if request.method == "POST":

        print("🔥 FORM SUBMITTED")

        name = request.POST.get("name")
        email = request.POST.get("email")
        phone = request.POST.get("phone")
        message = request.POST.get("message")

        Consultation.objects.create(
            name=name, email=email, phone=phone, message=message
        )

        return redirect("/en/?success=1")

    return render(
        request,
        "home.html",
        {"premium_services": _premium_services_queryset()},
    )


from .models import Booking


def my_assignments(request):
    bookings = Booking.objects.all().order_by("-created_at")

    return render(request, "employee_bookings.html", {"bookings": bookings})


from django.shortcuts import get_object_or_404, redirect
from .models import Booking


def toggle_artist_status(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)

    if request.method == "POST":
        booking.artist_status = not booking.artist_status
        booking.save()

    return redirect("my_assignments")


from django.utils import timezone
from datetime import timedelta


DEFAULT_PREMIUM_SERVICES = [
    {"title": "3D Wall Craft - Walls That Come Alive", "description": "Turn plain walls into extraordinary 3D masterpieces. Our creative wall art adds depth, style, and a striking visual impact to any space. Perfect for interiors that deserve a bold artistic touch.", "image_url": "/static/gallery/3d.jpeg", "book_slug": "3d-wall-art", "explore_slug": "3d-wall-art"},
    {"title": "3D Floor Magic - Walk on Art", "description": "Transform ordinary floors into stunning 3D visual experiences. Our realistic designs create breathtaking illusions that instantly capture attention. Perfect for homes, malls, offices, and commercial spaces.", "image_url": "/static/gallery/3dfloorsrt.jpeg", "book_slug": "3d-floor-art", "explore_slug": "3d-floor-art"},
    {"title": "Mural Works - Walls that Tell Your Story", "description": "Bring life and personality to your walls with custom mural designs. Our artistic murals transform blank spaces into meaningful visual stories. Perfect for homes, offices, and creative environments.", "image_url": "/static/gallery/m.jpeg", "book_slug": "mural", "explore_slug": "mural"},
    {"title": "Metro & Piller Art - Be Seen by Thousands", "description": "Align your brand with high-traffic metro spaces using bold visuals to reach daily commuters. Stay visible on the move and create a lasting impression that keeps your brand top-of-mind.", "image_url": "/static/gallery/ad.jpeg", "book_slug": "metro-advertisement", "explore_slug": "metro-advertisement"},
    {"title": "Urban Art - Make Your Brand Unmissable", "description": "Stand out in busy streets and public spaces with powerful outdoor advertising art. Our bold designs attract attention from a distance. Perfect for brands that want strong visibility.", "image_url": "/static/gallery/ambujaadvertisement.jpeg", "book_slug": "outdoor-advertisement", "explore_slug": "outdoor-advertisement"},
    {"title": "School Art - Walls That Inspire Young Minds", "description": "Transform school walls into colorful and educational spaces. Our creative artwork encourages imagination and curiosity in children. Perfect for making learning environments vibrant and engaging.", "image_url": "/static/gallery/school.jpg", "book_slug": "school-painting", "explore_slug": "school-painting"},
    {"title": "Selfi Spot - Turn Your Photo into Wall Art", "description": "Convert your favorite photos into stunning artistic portraits. Our custom selfie art transforms memories into beautiful wall decor. Perfect for homes, gifts, and personal spaces.", "image_url": "/static/gallery/selfie.jpeg", "book_slug": "selfie-painting", "explore_slug": "selfie-painting"},
    {"title": "Madhubani Art - Traditional Art for Modern Walls", "description": "Experience the beauty of authentic Madhubani art on your walls. Our hand-crafted designs bring rich culture and vibrant colors to your space. Perfect for elegant and traditional decor.", "image_url": "/static/gallery/Madhubani painting.jpeg", "book_slug": "madhubani-painting", "explore_slug": "madhubani-painting"},
    {"title": "Texture Wall - Luxury You Can Feel on Your Walls", "description": "Add depth and sophistication with modern texture art designs. Our premium finishes create stylish and luxurious interiors. Perfect for contemporary homes and designer spaces.", "image_url": "/static/gallery/Texture panting.jpeg", "book_slug": "texture-painting", "explore_slug": "texture-painting"},
    {"title": "Statue Craft - Masterpieces in Stone", "description": "Bring your spaces to life with premium hand-carved stone statues. Our custom sculptures add artistic depth, elegance, and a touch of heritage to gardens, corporate offices, and private properties.", "image_url": "/static/gallery/buddha_statue.jpg", "book_slug": "stone-murti", "explore_slug": "stone-murti"},
    {"title": "Scrap Craft - Art Made from Creativity & Recycling", "description": "Unique animal sculptures created from recycled materials. Our scrap art adds creativity and sustainability to parks, gardens, and public spaces. Perfect for eco-friendly artistic installations.", "image_url": "/static/gallery/Scrap animal.jpeg", "book_slug": "scrap-animal-art", "explore_slug": "scrap-animal-art"},
    {"title": "Tank Art - From Tank to Landmark", "description": "Turn ordinary water tanks into colorful artistic highlights. Our creative designs transform dull structures into vibrant visual attractions. Perfect for communities and public spaces.", "image_url": "/static/gallery/Water tank.jpeg", "book_slug": "nature-fountain", "explore_slug": "nature-fountain"},
    {"title": "Zenith Collection - Walls Full of Creativity and Life", "description": "Bring your walls to life with stunning statue-inspired designs. Our artwork adds depth, creativity, and a unique 3D feel to your space. Perfect for homes, gardens, and commercial areas that need a bold artistic touch.", "image_url": "/static/gallery/z.jpeg", "book_slug": "Zenith-Collection", "explore_slug": "Zenith-Collection"},
    {"title": "Cartoon Craft - Walls Full of Fun and Imagination", "description": "Bring joy and creativity with playful cartoon wall art. Our designs create lively spaces for kids rooms, schools, and play areas. Perfect for bright and cheerful environments.", "image_url": "/static/gallery/child.jpeg", "book_slug": "cartoon-painting", "explore_slug": "cartoon-painting"},
    {"title": "Home Decore Art - Make Your Home a Work of Art", "description": "Elevate your space with bespoke wall paintings and canvas art that add luxury and uniqueness. Expertly crafted by professional artists, each piece is a perfect masterpiece for your home.", "image_url": "/static/gallery/living.jpeg", "book_slug": "home-painting", "explore_slug": "home-painting"},
]


def _premium_services_queryset():
    from .models import PremiumService
    try:
        qs = PremiumService.objects.filter(is_active=True).order_by("sort_order", "id")
        if qs.exists():
            return list(qs)
        # Seed defaults once so they become editable from UI
        for idx, item in enumerate(DEFAULT_PREMIUM_SERVICES, start=1):
            PremiumService.objects.create(
                title=item["title"],
                description=item["description"],
                image_url=item["image_url"],
                book_slug=item["book_slug"],
                explore_slug=item["explore_slug"],
                sort_order=idx,
                is_active=True,
            )
        return list(PremiumService.objects.filter(is_active=True).order_by("sort_order", "id"))
    except (ProgrammingError, OperationalError):
        return DEFAULT_PREMIUM_SERVICES


def premium_services_editor(request):
    from .models import PremiumService

    _require_staff_or_404(request)
    _premium_services_queryset()

    if request.method == "POST":
        ids = request.POST.getlist("service_id[]")
        titles = request.POST.getlist("title[]")
        descriptions = request.POST.getlist("description[]")
        images = request.POST.getlist("image_url[]")
        books = request.POST.getlist("book_slug[]")
        explores = request.POST.getlist("explore_slug[]")
        active_ids = set(request.POST.getlist("active_ids[]"))
        remove_ids = set([x for x in request.POST.getlist("remove_ids[]") if x])

        for i in range(len(titles)):
            title = (titles[i] or "").strip()
            if not title:
                continue
            row_id = (ids[i] or "").strip() if i < len(ids) else ""
            if row_id and row_id in remove_ids:
                continue
            payload = {
                "title": title,
                "description": (descriptions[i] if i < len(descriptions) else "").strip(),
                "image_url": (images[i] if i < len(images) else "").strip(),
                "book_slug": (books[i] if i < len(books) else "").strip(),
                "explore_slug": (explores[i] if i < len(explores) else "").strip(),
                "sort_order": i + 1,
                "is_active": row_id in active_ids if row_id else True,
            }
            if row_id:
                PremiumService.objects.filter(id=row_id).update(**payload)
            else:
                PremiumService.objects.create(**payload)

        if remove_ids:
            PremiumService.objects.filter(id__in=list(remove_ids)).delete()

        messages.success(request, "Premium services updated successfully.")
        return redirect("premium_services_editor")

    try:
        services = PremiumService.objects.all().order_by("sort_order", "id")
    except (ProgrammingError, OperationalError):
        services = []
    return render(request, "premium_services_editor.html", {"services": services})


def my_bookings(request):
    bookings = Booking.objects.filter(customer_user=request.user)

    for b in bookings:
        time_diff = timezone.now() - b.created
        b.can_cancel = time_diff <= timedelta(hours=2)

    return render(request, "my_bookings.html", {"bookings": bookings})


from django.utils import timezone
from datetime import timedelta
from django.contrib import messages


@login_required
def cancel_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)

    # 🔒 only owner
    if booking.customer_user_id != request.user.id:
        return redirect("my_orders")

    # ⏱ time check
    now = timezone.now()
    time_diff = now - booking.created  # ✅ FIXED

    # ❌ after 2 hours
    if time_diff > timedelta(hours=2):
        messages.error(
            request, "❌ Order is already processed. Please contact support to cancel."
        )
        return redirect("my_orders")

    # ❌ completed
    if booking.status == "completed":
        messages.error(request, "❌ Completed order cannot be cancelled!")
        return redirect("my_orders")

    # ❌ already cancelled
    if booking.status == "cancelled":
        messages.warning(request, "⚠️ Order already cancelled.")
        return redirect("my_orders")

    # ✅ cancel
    booking.status = "cancelled"
    booking.save()

    messages.success(request, "✅ Order cancelled successfully!")
    return redirect("/my-orders/?cancelled=true")
